from openpyxl import load_workbook
from openpyxl import Workbook

#Lista de arquivos Excel para ser consolidada
lista_arquivos = ['CustosAutom', 'PopulacaoPOA','SuperMercado']

#Novo arquivo
wb = Workbook()
nome_arquivo_final = 'resultado.xlsx'

for nome_arquivo in lista_arquivos:
    print('Lendo arquivo %s...'%nome_arquivo)
    arquivo = load_workbook(filename='%s.xlsx'%nome_arquivo)
    sheet = arquivo[nome_arquivo]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column
    
    ws = wb.create_sheet(title=nome_arquivo)
    
#Passar os dados de um arquivo para outro

for i in range(1,max_linhas+1):
    for j in range(1,max_colunas + 1 ):
        c = sheet.cell(row=i,column=j)

        ws.cell(row=i,column=j).value = c.value


print('Criando arquivo final %s...'%nome_arquivo)
wb.remove(wb['Sheet'])
wb.save(nome_arquivo_final)