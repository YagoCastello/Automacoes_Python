from openpyxl import load_workbook

# Carregar o arquivo Excel
wb = load_workbook(filename='nomes.xlsx')

# Selecionar a planilha desejada
planilha = wb['Planilha1']

# Inicializar a variável para somar as idades
soma_idade = 0

# Iterar pelas linhas 2 a 4 (inclusive)
for i in range(2, 5):
    nome = planilha[f'A{i}'].value
    idade = planilha[f'B{i}'].value
    
    # Exibir o nome e a idade
    print(f'{nome} tem {idade} anos.')
    
    # Somar a idade
    soma_idade += int(idade)

# Inserir a soma das idades na célula B5
planilha['A5'] = 'soma_idade'

planilha['B5'] = '=SUM(B2:B4)'

planilha['A7'] = 'ALUNOS'
planilha.merge_cells('A7:B7')
# planilha.unmerge_cells('A7:B7')


# Salvar as alterações no arquivo Excel
wb.save(filename='nomes.xlsx')
